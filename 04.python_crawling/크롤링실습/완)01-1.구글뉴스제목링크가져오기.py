from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

# 크롬 드라이버 자동 업데이트
from webdriver_manager.chrome import ChromeDriverManager

import time
import pyautogui
import openpyxl

# 브라우저 꺼짐 방지
chrome_options = Options()
chrome_options.add_experimental_option("detach",True)

# 불필요한 에러 메시지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

# 엑셀 파일 로드
fpath=r'C:\Users\lg\Desktop\startcoding\04.python_crawling\크롤링실습\02.파이썬엑셀다루기\크롤링 자료 엑셀 정리.xlsx'
wb = openpyxl.load_workbook(fpath)

ws = wb.active

# 웹페이지 설정, 최대화
driver.implicitly_wait(10) # 웹 페이지가 로딩 될 때까지 10초는 기다림
driver.maximize_window() #화면 최대화

# 웹페이지 해당 주소 이동
keyword = pyautogui.prompt("검색어를 입력하세요>>>")
lastpage = pyautogui.prompt("마지막 페이지번호를 입력해 주세요")
pageNum = 1
row=1
for i in range(0, int(lastpage) * 10, 10):
    print(f"{pageNum}페이지 입니다. =============================")
    driver.get(f"https://www.google.com/search?q={keyword}&tbm=nws&ei=Rl-PY9_4O8qCoASVqJCIDQ&start={i}&sa=N&ved=2ahUKEwjfs8P8puX7AhVKAYgKHRUUBNE4HhDy0wN6BAgBEAQ&biw=1536&bih=708&dpr=1.25")
    
# 기사제목
    # names = driver.find_elements(By.CSS_SELECTOR, ".mCBkyc")
    # for name in names:
    #     print(name.text)
    #     # ws[f'A{row}'] = name.text
    #     # ws[f'B{row}'] = href
    #     row=row+1
    
    hrefs = driver.find_elements(By.CSS_SELECTOR, ".YKoRaf")
    for href in hrefs:
        print(href)
        # ws[f'A{row}'] = name.text
        # ws[f'B{row}'] = href
        row=row+1
    pageNum = pageNum + 1

wb.save(fpath)
 